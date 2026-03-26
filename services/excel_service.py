import os
from openpyxl import Workbook, load_workbook

EXCEL_FILE = "registration.xlsx"

def save_registration_to_excel(full_name, email, phone, role):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Registrations"
        ws.append(["Full Name", "Email", "Phone", "Role"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Registrations"]
    ws.append([full_name, email, phone, role])
    wb.save(EXCEL_FILE)